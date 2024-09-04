
def InputDataToSubject(workbook,sheet,row,value):
    sheet = workbook[str(sheet)]
    num = 7
    while True:
        cell_value = sheet[row+str(num)]
        if cell_value.value is None:
            cell_value.value = value
            #print(f"Sum Value {num-3} = {cell_value.value}")
            break
        else:
            num +=1
    sheet['A'+str(num)].value = num - 6

def InputDataToSum(workbook,sheet,row,value):
    sheet = workbook[str(sheet)]
    num = 4
    while True:
        cell_value = sheet[row+str(num)]
        if cell_value.value is None:
            cell_value.value = value
            #print(f"Sum Value {num-3} = {cell_value.value}")
            break
        else:
            num +=1
    sheet['A'+str(num)].value = num - 3

def InputDataToIncome(workbook,sheet,row,value):
    sheet = workbook[str(sheet)]
    num = 3
    sheet[row+str(num)].value = value

from openpyxl import load_workbook
#สำหรับหาผลรวมของแต่ละแถวแนวตั้งของค่า test01       
def columnSum(workbook, sheet_name, column):
    wb = load_workbook(workbook)
    ws = wb[sheet_name]

    total_sum = 0
    for row in range(7, ws.max_row + 1):
        
        cell = ws[column + str(row)].value
        total_sum += cell        

    return total_sum

def TestMergeCell(filename,sheet_name,cell):
    wb = load_workbook(filename)
    sheet = wb[sheet_name]
    c = sheet[cell]
    for mergedCell in sheet.merged_cells.ranges:
        if (c.coordinate in mergedCell):
            return True
    return False