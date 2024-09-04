from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
sheet = wb.active
SubjectID = {
            501171:'การเรียนรู้เพื่อการแลกเปลี่ยนในพยาบาล',
            501172:'มโนนิติ ทฤษฎีและวิชาชีพการพยาบาล',
            501005:'การสร้างเสริมสุขภาพและการดูเบื้องต้น',
            501251:'การพยาบาลสุขภาพจิต',
            501261:'การประเมินภาวะสุขภาพ',
            501004:'ภาษาอังกฤษเชิงวิชาชีพสำหรับพยาบาล',
            501221:'การพยาบาลเด็ก 1',
            501231:'การพยาบาลผู้ใหญ่ 1',
            501232:'การพยาบาลผู้ใหญ่ 2',
            501273:'หลักการและเทคนิคทางการพยาบาล',
            501274:'กฎหมายและจริยศาสตร์สำหรับพยาบาล',
            501233:'ปฏิบัติการพยาบาลผู้ใหญ่',
            501252:'ปฏิบัติการพยาบาลสุขภาพจิต',
            501275:'ปฏิบัติการหลักการและเทคนิคการพยาบาล',
            501311:'การพยาบาลมารดา ทารกและการผดุงครรภ์ 1',
            501322:'การพยาบาลเด็ก 2',
            501334:'การพยาบาลผู้ใหญ่ 3',
            501341:'การพยาบาลผู้สูงอายุ',
            501353:'การพยาบาลจิตเวช',
            501362:'การพยาบาลชุมชน 1',
            501363:'การรักษาพยาบาลเบื้องต้น',
            501391:'วิทยานิพนธ์ระดับปริญญาตรี 1',
            501323:'ปฏิบัติการพยาบาลเด็ก 1',
            501335:'ปฏิบัติการพยาบาลผู้ใหญ่ 2',
            501342:'ปฏิบัติการพยาบาลผู้สูงอายุ 1',
            501354:'ปฏิบัติการพยาบาลจิตเวช',
            501392:'วิทยานิพนธ์ระดับปริญญาตรี 2',
            501312:'ปฏิบัติการพยาบาลมารดา ทารกและการผดุงครรภ์ 1',
            501324:'ปฏิบัติการพยาบาลเด็ก 2',
            501413:'ปฏิบัติการพยาบาลมารดา ทารกและการผดุงครรภ์ 2',
            501435:'การพยาบาลผู้และผู้สูงอายุ 3',
            501436:'ปฏิบัติการพยาบาลผู้และผู้สูงอายุ 3',
            501463:'ปฏิบัติการรักษาพยาบาลเบื้องต้น',
            501464:'การพยาบาลอนามัยชุมชน 2',
            501480:'การจัดการทางการพยาบาล',
            501481:'ปฏิบัติการการจัดการทางการพยาบาล',
            501414:'ปฏิบัติการพยาบาลมารดา ทารกและการผดุงครรภ์ 2',
            501465:'ปฏิบัติการพยาบาลอนามัยชุมชน',
            501482:'สัมมนาประเด็นและแนวโน้มของวิชาชีพพยาบาล',
            501483:'ปฏิบัติการพยาบาลวิชาชีพในสาขาที่เลือกสรร'
        }

def createSubjectSheet(filename,ID):
    sheet.merge_cells('A1:D1')
    cell = sheet['A1']
    cell.value = f'สรุปรายการเบิกเงิน รายวิชา {ID} ...' ##
    cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells('A2:D2')
    cell = sheet['A2'] 
    cell.value = 'ภาคเรียน ...   ปีการศึกษา  ...'##
    cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells('A3:D3')
    cell = sheet['A3']  
    cell.value = 'ผู้รับผิดชอบรายวิชา  : ...'  ##
    cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('A4:A6')
    cell = sheet['A4']  
    cell.value = 'ที่'  

    sheet.merge_cells('B4:B6')
    cell = sheet['B4']  
    cell.value = 'วัน/เดือน/ปี'  
    cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('C4:C6')
    cell = sheet['C4']  
    cell.value = 'อ้างอิง'  
    cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('D4:D6')
    cell = sheet['D4']  
    cell.value = 'รายการ/ชื่อ-สกุล'  
    cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('E4:I4')
    cell = sheet['E4']  
    cell.value = 'รายการเบิก'  
    sheet.merge_cells('E5:G5')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell = sheet['E5']  
    cell.value = 'ค่าตอบแทน'  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell = sheet['E6']  
    cell.value = 'ค่าสอน'  
    cell = sheet['F6']  
    cell.value = 'แหล่งฝึก'  
    cell = sheet['G6']  
    cell.value = 'ปฐมนิเทศ'  
    cell = sheet['H5']  
    cell.value = 'ค่าใช้สอย'  
    cell = sheet['H6']  
    cell.value = 'ค่าจ้างเหมารถ'  
    cell = sheet['I5']  
    cell.value = 'ค่าวัสดุ'  
    cell = sheet['I6']  
    cell.value = 'น้ำมันเชื้อเพลิง'

    sheet.merge_cells('J4:L4')
    cell = sheet['J4']  
    cell.value = 'รายการเบิก-คณาจารย์'  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells('J5:L5')
    cell = sheet['J5']  
    cell.value = 'ใช้สอย'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell = sheet['J6']  
    cell.value = 'ค่าเบี้ยเลี้ยง'
    cell = sheet['K6']  
    cell.value = 'ค่าที่พัก'
    cell = sheet['L6']  
    cell.value = 'ค่าเดินทาง'

    sheet.merge_cells('M4:N4')
    cell = sheet['M4']  
    cell.value = 'รายการเบิก-นิสิต'  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells('M5:N5')
    cell = sheet['M5']  
    cell.value = 'ใช้สอย'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell = sheet['M6']  
    cell.value = 'ค่าที่พัก'
    cell = sheet['N6']  
    cell.value = 'ค่าเดินทาง'

    sheet.merge_cells('O4:Q4')
    cell = sheet['O4']  
    cell.value = 'รายการเบิก-พนักงานขับรถ'  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells('O5:Q5')
    cell = sheet['O5']  
    cell.value = 'ใช้สอย'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell = sheet['O6']  
    cell.value = 'เบี้ยเลี้ยง'
    cell = sheet['P6']  
    cell.value = 'OT'
    cell = sheet['Q6']  
    cell.value = 'ค่าโบนัสพิเศษ'
    sheet.merge_cells('R4:R6')
    cell = sheet['R4']
    cell.value = 'รวมรายการเบิก'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells('S4:S6')
    cell = sheet['S4']
    cell.value = 'หมายเหตุ'
    cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('A7:D7')
    cell = sheet['A7']
    cell.value = 'รวม'
    cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(filename)  

createSubjectSheet('Test01.xlsx',501004)
for i in SubjectID:
    sheet.merge_cells('A1:D1')
    cell = sheet['A1']
    cell.value = f'สรุปรายการเบิกเงิน รายวิชา {i} {SubjectID[i]}' ##
    cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.copy_worksheet(wb['Sheet'])
    CopyNameSheet = wb['Sheet Copy']
    CopyNameSheet.title = str(i)
wb.save('Test01.xlsx')  

def createSumSheet(filename):

    sheet.merge_cells('A1:A2')
    cell = sheet['A1']  
    cell.value = 'ที่'  

    sheet.merge_cells('B1:B2')
    cell = sheet['B1']
    cell.value = 'รหัสวิชา'  
    cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('C1:C2')
    cell = sheet['C1']  
    cell.value = 'วัน/เดือน/ปี'  
    cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('D1:D2')
    cell = sheet['D1']  
    cell.value = 'อ้างอิง'  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells('E1:E2')
    cell = sheet['E1']  
    cell.value = 'รายการ/ชื่อ-สกุล'  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells('F1:G2')
    cell = sheet['F1']  
    cell.value = 'โครงการจัดการเรียนการสอน ป.ตรี'  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells('A3:E3')
    cell = sheet['A3']  
    cell.value = 'งบประมาณที่ได้รับการจัดสรร'  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    cell = sheet['F2']
    cell.value = 'แผ่นดิน'
    
    cell = sheet['G2']
    cell.value = 'รายได้'

    wb.save(filename)
# createSumSheet('TestSum01.xlsx')
# wb.save('TestSum01.xlsx')  