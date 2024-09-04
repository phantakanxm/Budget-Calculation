from tkinter import Tk, Canvas , Button
from tkinter import messagebox
from tkinter import *  #ดึงหน้า gui
from tkinter import ttk 
from tkcalendar import DateEntry #ดึงปฏิทิน

from pdf import *

from Sumwindow import * #เปิดหน้าแสดงผลรวม
from InputData import * #function ส่งข้อมูลไปยัง excel

from openpyxl import load_workbook #ดึงข้อมูลจากไฟล์ excel

import sys
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
#เพื่อเรียกใช้ dictionary

filenameID = 'Test01.xlsx' #ชื่อไฟล์ที่ลงข้อมูล
filenameSum = 'งบประมาณรวม.xlsx' #ชื่อไฟล์ที่ลงข้อมูลรวม

class SubjectDataEntryApp:
    def __init__(self, root):
        
        self.root = root
        self.root.geometry("400x300")
        self.root.configure(bg="#FFFFFF")
        
        self.ascii = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        
        #กำหนดรหัสวิชา : ชื่อวิชา
        self.SubjectID = {
            501171:'การเรียนรู้เพื่อการแลกเปลี่ยนในพยาบาล',
            501172:'มโนนิติ ทฤษฎีและวิชาชีพการพยาบาล',
            501005:'การสร้างเสริมสุขภาพและการดูเบื้องต้น',
            501251:'การพยาบาลสุขภาพจิต',
            501261:'การประเมินภาวะสุขภาพ',
            501004:'ภาษาอังกฤษเชิงวิชาชีพสำหรับพยาบาล',
            501221:'การพยาบาลเด็ก 1',
            501231:'การพยาบาลผู้ใหญ่ 1',
            501232:'การพยาบาลผู้ใหญ่ 2',
            501273:'หลักการและเทคนิคทางการพยาบาล',
            501274:'กฎหมายและจริยศาสตร์สำหรับพยาบาล',
            501233:'ปฏิบัติการพยาบาลผู้ใหญ่',
            501252:'ปฏิบัติการพยาบาลสุขภาพจิต',
            501275:'ปฏิบัติการหลักการและเทคนิคการพยาบาล',
            501311:'การพยาบาลมารดา ทารกและการผดุงครรภ์ 1',
            501322:'การพยาบาลเด็ก 2',
            501334:'การพยาบาลผู้ใหญ่ 3',
            501341:'การพยาบาลผู้สูงอายุ',
            501353:'การพยาบาลจิตเวช',
            501362:'การพยาบาลชุมชน 1',
            501363:'การรักษาพยาบาลเบื้องต้น',
            501391:'วิทยานิพนธ์ระดับปริญญาตรี 1',
            501323:'ปฏิบัติการพยาบาลเด็ก 1',
            501335:'ปฏิบัติการพยาบาลผู้ใหญ่ 2',
            501342:'ปฏิบัติการพยาบาลผู้สูงอายุ 1',
            501354:'ปฏิบัติการพยาบาลจิตเวช',
            501392:'วิทยานิพนธ์ระดับปริญญาตรี 2',
            501312:'ปฏิบัติการพยาบาลมารดา ทารกและการผดุงครรภ์ 1',
            501324:'ปฏิบัติการพยาบาลเด็ก 2',
            501413:'ปฏิบัติการพยาบาลมารดา ทารกและการผดุงครรภ์ 2',
            501435:'การพยาบาลผู้และผู้สูงอายุ 3',
            501436:'ปฏิบัติการพยาบาลผู้และผู้สูงอายุ 3',
            501463:'ปฏิบัติการรักษาพยาบาลเบื้องต้น',
            501464:'การพยาบาลอนามัยชุมชน 2',
            501480:'การจัดการทางการพยาบาล',
            501481:'ปฏิบัติการการจัดการทางการพยาบาล',
            501414:'ปฏิบัติการพยาบาลมารดา ทารกและการผดุงครรภ์ 2',
            501465:'ปฏิบัติการพยาบาลอนามัยชุมชน',
            501482:'สัมมนาประเด็นและแนวโน้มของวิชาชีพพยาบาล',
            501483:'ปฏิบัติการพยาบาลวิชาชีพในสาขาที่เลือกสรร'
        }
        
        self.txt = IntVar()
        self.canvas = Canvas(self.root, bg="#FFFFFF", height=300, width=400)
        self.canvas.place(x=0, y=0)
        self.canvas.create_text(168.0, 68.0, anchor="nw", text="ใส่รหัสรายวิชา", fill="#000000", font=("Inter", 12 * -1))
        #ช่องใส่รหัสวิชา
        
        self.entry_1 = ttk.Entry(root,style="TEntry",textvariable = self.txt)
        self.entry_1.place(x=133.0, y=93.0, width=141.0, height=25.0)
        
        #ปุ่มตกลง
        self.button_1 = Button(text = "ตกลง",borderwidth=0, highlightthickness=0, relief="flat", command=lambda:self.open_window_2(self.txt.get()))
        self.button_1.place(x=175.0, y=135.0, width=50.0, height=23.0)
        
        #ปุ่มใส่งบประมาณใหม่
        self.button_2 = Button(text = "ใส่งบประมาณ",borderwidth=0, highlightthickness=0, relief="flat", command= lambda:self.open_window_3())
        self.button_2.place(x=75.0, y=200.0, width=100.0, height=23.0)
        
        #ปุ่มไปหาหน้าแสดงผลรวม
        self.button_3 = Button(text = "SUM",borderwidth=0, highlightthickness=0, relief="flat", command= lambda:open_window_4())
        self.button_3.place(x=225.0, y=200.0, width=100.0, height=23.0)
        
    def open_window_3(self) :
        self.window3 = Tk()
        self.window3.geometry("600x450")
        self.window3.configure(bg="#FFFFFF")
        
        canvas = Canvas(self.window3,bg = "#FFFFFF",height = 760,width = 900,bd = 0,highlightthickness = 0,relief = "ridge")
        canvas.place(x = 0, y = 0)
        canvas.create_rectangle(-1.0,1011.0,1920.0093994140625,1012.0,fill="#000000",outline="")
        
        
        m1 = IntVar()
        canvas.create_text(245.0,100.0,anchor="nw",text="กรุณาระบุจำนวณเงิน",fill="#000000",font=("Inter SemiBold", 15 * -1))
        entry_m1 = ttk.Entry(self.window3,style="TEntry",textvariable = m1, )
        entry_m1.place(x=220.0,y=150.0,width=165.0,height=26.0)
        
        choice = StringVar(self.window3, value = "โปรดเลือกงบประมาณ")
        combo = ttk.Combobox(self.window3, width = 20, textvariable=choice)
        combo["value"] = ("งบประมาณแผ่นดิน","งบประมาณรายได้")
        combo.place(x=230.0,y=200.0)
        
        def savemoney() :
            self.amout = entry_m1.get()
            currency = choice.get()
            
            workbookSum = load_workbook(filenameSum)
            workbookSum.active
                        
            if currency == "งบประมาณแผ่นดิน" :
                InputDataToIncome(workbookSum,'Sheet1','F',self.amout)
                messagebox.showinfo("Success","ส่งข้อมูลเรียบร้อย")
                self.window3.destroy()

            elif currency == "งบประมาณรายได้" :
                InputDataToIncome(workbookSum,'Sheet1','G',self.amout)
                messagebox.showinfo("Success","ส่งข้อมูลเรียบร้อย")
                self.window3.destroy()   
                     
            workbookSum.save(filenameSum)
        button1_m1 = Button(self.window3,text = " ส่งข้อมูล ",font=("Inter SemiBold", 15 * -1),command = savemoney).place(x=265,y=260)    
    def open_window_2(self,ID):
        try:
            if ID in self.SubjectID:
                
                window2 = Tk()
                window2.geometry("900x760")
                window2.configure(bg="#FFFFFF")

                canvas = Canvas(window2,bg = "#FFFFFF",height = 760,width = 900,bd = 0,highlightthickness = 0,relief = "ridge")

                canvas.place(x = 0, y = 0)
                canvas.create_rectangle(-1.0,1011.0,1920.0093994140625,1012.0,fill="#000000",outline="")
            
                x2 = StringVar()
                x3 = StringVar()
                x4 = StringVar()

                x5,x6,x7,x8,x9,x10 = IntVar(),IntVar(),IntVar(),IntVar(),IntVar(),IntVar()
                x11,x12,x13,x14,x15,x16,x17 = IntVar(),IntVar(),IntVar(),IntVar(),IntVar(),IntVar(),IntVar()

                #ชื่อหัว
                canvas.create_rectangle(9.0,21.0,370.0,58.0,fill="#EEEEEE",outline="")
                canvas.create_text(22.0,31.0,anchor="nw",text=f"{ID} {self.SubjectID[ID]}",fill="#000000",font=("Inter SemiBold", 15 * -1))
                canvas.create_rectangle(-0.998779296875,80.95834350585938,900.00341796875,81.95834350585938,fill="#000000",outline="")

                #กล่องป้อนข้อมูล วัน เดือน ปี
                canvas.create_text(17.0,112.0,anchor="nw",text="วัน/เดือน/ปี",fill="#000000",font=("Inter SemiBold", 15 * -1))
                cal=DateEntry(window2,selectmode='day',date_pattern="dd/mm/yyyy")
                cal.place(x=131.0,y=105.0,width=165.0,height=26.0)

                #กล่องป้อนข้อมูล อ้างอิง
                canvas.create_text(17.0,160.0,anchor="nw",text="เลขที่หนังสือ ",fill="#000000",font=("Inter SemiBold", 15 * -1))
                entry_3 = ttk.Entry(window2,style="TEntry",textvariable = x3)
                entry_3.place(x=131.0,y=154.0,width=165.0,height=26.0)

                #กล่องป้อนข้อมูล รายการชื่อ-สกุล
                canvas.create_text(17.0,209.0,anchor="nw",text="รายการ/ชื่อ-สกุล",fill="#000000",font=("Inter SemiBold", 15 * -1))
                entry_4 = ttk.Entry(window2,style="TEntry",textvariable = x4)
                entry_4.place(x=131.0,y=203.0,width=165.0,height=25.0)

                #รายการเบิก ค่าตอบแทน
                canvas.create_text(17.0,256.0,anchor="nw",text="รายการเบิก",fill="#000000",font=("Inter SemiBold", 18 * -1))
                canvas.create_text(131.0,262.0,anchor="nw",text="ค่าตอบแทน",fill="#000000",font=("Inter SemiBold", 15 * -1))
                #กล่องป้อนข้อมูล ค่าสอน
                entry_5 = ttk.Entry(window2,style="TEntry",textvariable = x5)
                entry_5.place(x=131.0,y=286.0,width=165.0,height=26.0)
                canvas.create_text(47.0,296.0,anchor="nw",text="ค่าสอน",fill="#000000",font=("Inter SemiBold", 15 * -1))
                #กล่องป้อนข้อมูล แหล่งฝึก
                entry_6 = ttk.Entry(window2,style="TEntry",textvariable = x6)
                entry_6.place(x=131.0,y=335.0,width=165.0,height=26.0)
                canvas.create_text(47.0,345.0,anchor="nw",text="แหล่งฝึก",fill="#000000",font=("Inter SemiBold", 15 * -1))
                #กล่องป้อนข้อมูล ปฐมนิเทศ
                entry_7 = ttk.Entry(window2,style="TEntry",textvariable = x7)
                entry_7.place(x=131.0,y=389.0,width=165.0,height=26.0)
                canvas.create_text(47.0,399.0,anchor="nw",text="ปฐมนิเทศ",fill="#000000",font=("Inter SemiBold", 15 * -1))

                #ค่าใช้สอย
                canvas.create_text(131.0,433.0,anchor="nw",text="ค่าใช้สอย",fill="#000000",font=("Inter SemiBold", 15 * -1))
                #กล่องป้อนข้อมูล ค่าจ้างเหมารถ
                entry_8 = ttk.Entry(window2,style="TEntry",textvariable = x8)
                entry_8.place(x=131.0,y=452.0,width=165.0,height=25.0)
                canvas.create_text(23.0,461.0,anchor="nw",text="ค่าจ้างเหมารถ",fill="#000000",font=("Inter SemiBold", 15 * -1))

                #ค่าวัสดุ
                canvas.create_text(131.0,501.0,anchor="nw",text="ค่าวัสดุ",fill="#000000",font=("Inter SemiBold", 15 * -1))
                #กล่องป้อนข้อมูล นํ้ามันเชื้อเพลิง
                entry_9 = ttk.Entry(window2,style="TEntry",textvariable = x9)
                entry_9.place(x=131.0,y=520.0,width=165.0,height=25.0)
                canvas.create_text(23.0,529.0,anchor="nw",text="นํ้ามันเชื้อเพลิง",fill="#000000",font=("Inter SemiBold", 15 * -1))

                #รายการเบิก - คณาอาจารย์ และ ใช้สอย
                canvas.create_text(386.0,206.0,anchor="nw",text="รายการเบิก-คณาอาจารย์",fill="#000000",font=("Inter SemiBold", 18 * -1))
                canvas.create_text(608.0,215.0,anchor="nw",text="ใช้สอย",fill="#000000",font=("Inter SemiBold", 15 * -1))
                #กล่องป้อนข้อมูล ค่าเบี้ยเลี้ยง
                entry_10 = ttk.Entry(window2,style="TEntry",textvariable = x10)
                entry_10.place(x=608.0,y=239.0,width=164.0,height=26.0)
                canvas.create_text(514.0,247.0,anchor="nw",text="ค่าเบี้ยเลี้ยง",fill="#000000",font=("Inter SemiBold", 15 * -1))
                #กล่องป้อนข้อมูล ค่าที่พัก
                entry_11 = ttk.Entry(window2,style="TEntry",textvariable = x11)
                entry_11.place(x=608.0,y=288.0,width=164.0,height=26.0)
                canvas.create_text(514.0,294.0,anchor="nw",text="ค่าที่พัก",fill="#000000",font=("Inter SemiBold", 15 * -1))
                #กล่องป้อนข้อมูล ค่าเดินทาง
                entry_12 = ttk.Entry(window2,style="TEntry",textvariable = x12)
                entry_12.place(x=608.0,y=342.0,width=164.0,height=26.0)
                canvas.create_text(514.0,350.0,anchor="nw",text="ค่าเดินทาง",fill="#000000",font=("Inter SemiBold", 15 * -1))

                #ข้อความ รายการเบิกนิสิต และ ใช้สอย
                canvas.create_text(447.0,384.0,anchor="nw",text="รายการเบิก-นิสิต",fill="#000000",font=("Inter SemiBold", 18 * -1))
                canvas.create_text(608.0,393.0,anchor="nw",text="ใช้สอย",fill="#000000",font=("Inter SemiBold", 15 * -1))
                #กล่องป้อนข้อมูล ค่าที่พัก
                entry_13 = ttk.Entry(window2,style="TEntry",textvariable = x13)
                entry_13.place(x=608.0,y=416.0,width=164.0,height=25.0)
                canvas.create_text(514.0,424.0,anchor="nw",text="ค่าที่พัก",fill="#000000",font=("Inter SemiBold", 15 * -1))
                #กล่องป้อนข้อมูล ค่าเดินทาง
                entry_14 = ttk.Entry(window2,style="TEntry",textvariable = x14)
                entry_14.place(x=608.0,y=464.0,width=164.0,height=26.0)
                canvas.create_text(514.0,473.0,anchor="nw",text="ค่าเดินทาง",fill="#000000",font=("Inter SemiBold", 15 * -1))

                #ข้อความ รายการเบิก-พนักงานขับรถ และ ใช้สอย
                canvas.create_text(368.0,510.0,anchor="nw",text="รายการเบิก-พนักงานขักรถ",fill="#000000",font=("Inter SemiBold", 18 * -1))
                canvas.create_text(608.0,519.0,anchor="nw",text="ใช้สอย",fill="#000000",font=("Inter SemiBold", 15 * -1))
                #กล่องป้อนข้อมูล เบี้ยเลี้ยง
                entry_15 = ttk.Entry(window2,style="TEntry",textvariable = x15)
                entry_15.place(x=608.0,y=543.0,width=164.0,height=25.0)
                canvas.create_text(521.0,551.0,anchor="nw",text="เบี้ยเลี้ยง",fill="#000000",font=("Inter SemiBold", 15 * -1))
                #กล่องป้อนข้อมูล OT
                entry_16=ttk.Entry(window2,style="TEntry",textvariable = x16)
                entry_16.place(x=608.0,y=592.0,width=164.0,height=25.0)
                canvas.create_text(528.0,599.0,anchor="nw",text="OT",fill="#000000",font=("Inter SemiBold", 15 * -1))
                #กล่องป้อนข้อมูล ค่าโบนัสพิเศษ
                entry_17=ttk.Entry(window2,style="TEntry",textvariable = x17)
                entry_17.place(x=608.0,y=641.0,width=164.0,height=25.0)
                canvas.create_text(521.0,641.0,anchor="nw",text="ค่าโบนัสพิเศษ",fill="#000000",font=("Inter SemiBold", 15 * -1))
                
                # สร้างปุ่ม Attach File
                attach_button1 = tk.Button(window2, text="Attach File", command=lambda:attach_file(file_label,attach_button1))
                attach_button1.place(x=225,y=575)
                file_label = tk.Label(window2, text="")
                file_label.place(x=225,y=575)

                
                #กล่องเลือกงบประมาณ
                choicew2 = StringVar(window2, value = "โปรดเลือกงบประมาณ")
                combow2 = ttk.Combobox(window2, width = 20, textvariable=choicew2)
                combow2["value"] = ("งบประมาณแผ่นดิน","งบประมาณรายได้")
                combow2.place(x=50.0,y=580.0)
                
                window2.after(100, lambda:combow2.set("โปรดเลือกงบประมาณ"))
                sheet_name = str(ID) #แทนชื่อชีทด้วยรหัสวิชา
                
                def success():
                    self.ValueList = list()
                    self.SumValueList = list()
                    confirm = messagebox.askquestion("ยืนยัน","คุณต้องส่งข้อมูลหรือไม่" )
                    if confirm == "yes" :
                        budgetChoice = combow2.get()
                        if budgetChoice not in ("งบประมาณแผ่นดิน","งบประมาณรายได้"):
                            messagebox.showinfo("ข้อผิดพลาด","กรุณาเลือกงบประมาณ")
                            return
                        
                        dt = cal.get_date()
                        dt3=dt.strftime("%d/%m/%Y")
                        self.x2_value = dt3 if dt3 else ''
                        self.ValueList.append(self.x2_value)
                        
                        self.x3_value = entry_3.get() if entry_3.get() else ''
                        self.ValueList.append(self.x3_value)
                        
                        self.x4_value = entry_4.get() if entry_4.get() else ''
                        self.ValueList.append(self.x4_value)
                        
                        self.x5_value = int(entry_5.get()) if entry_5.get() else 0
                        self.ValueList.append(self.x5_value)
                        self.SumValueList.append(self.x5_value)
                        
                        self.x6_value = int(entry_6.get()) if entry_6.get() else 0
                        self.ValueList.append(self.x6_value)
                        self.SumValueList.append(self.x6_value)
                        
                        self.x7_value = int(entry_7.get()) if entry_7.get() else 0
                        self.ValueList.append(self.x7_value)
                        self.SumValueList.append(self.x7_value)
                        
                        self.x8_value = int(entry_8.get()) if entry_8.get() else 0
                        self.ValueList.append(self.x8_value)
                        self.SumValueList.append(self.x8_value)
                        
                        self.x9_value = int(entry_9.get()) if entry_9.get() else 0
                        self.ValueList.append(self.x9_value)
                        self.SumValueList.append(self.x9_value)
                        
                        self.x10_value = int(entry_10.get()) if entry_10.get() else 0
                        self.ValueList.append(self.x10_value)
                        self.SumValueList.append(self.x10_value)
                        
                        self.x11_value = int(entry_11.get()) if entry_11.get() else 0
                        self.ValueList.append(self.x11_value)
                        self.SumValueList.append(self.x11_value)
                        
                        self.x12_value = int(entry_12.get()) if entry_12.get() else 0
                        self.ValueList.append(self.x12_value)
                        self.SumValueList.append(self.x12_value)
                        
                        self.x13_value = int(entry_13.get()) if entry_13.get() else 0
                        self.ValueList.append(self.x13_value)
                        self.SumValueList.append(self.x13_value)
                        
                        self.x14_value = int(entry_14.get()) if entry_14.get() else 0
                        self.ValueList.append(self.x14_value)
                        self.SumValueList.append(self.x14_value)
                        
                        self.x15_value = int(entry_15.get()) if entry_15.get() else 0
                        self.ValueList.append(self.x15_value)
                        self.SumValueList.append(self.x15_value)
                        
                        self.x16_value = int(entry_16.get()) if entry_16.get() else 0
                        self.ValueList.append(self.x16_value)
                        self.SumValueList.append(self.x16_value)
                        
                        self.x17_value = int(entry_17.get()) if entry_17.get() else 0
                        self.ValueList.append(self.x17_value)
                        self.SumValueList.append(self.x17_value)

                        self.sum = 0
                        for i in self.SumValueList:
                            self.sum += i
                        self.ValueList.append(self.sum)
                        
                        workbookID = load_workbook(filenameID)
                        workbookID.active
                        sheet = workbookID[sheet_name]
                        num = sheet.max_row

                        sheet.unmerge_cells(f'A{num}:D{num}')
                        sheet.merge_cells(f'A{num+1}:D{num+1}')
                            
                        for j in range(len(self.ValueList)):
                            print(f"inputdata {self.ascii[j]} : {self.ValueList[j]}")
                            InputDataToSubject(workbookID,sheet_name,self.ascii[j+1],self.ValueList[j])
                        
                        cell = sheet[f'A{num+1}']  
                        cell.value = 'รวม'
                        for row in range(4,18):
                            print(f"column : {self.ascii[row]}")
                            Csum = columnSum(filenameID,sheet_name,self.ascii[row])
                            cell = sheet[f'{self.ascii[row]}{num+1}']
                            cell.value = Csum

                        workbookID.save(filenameID)
                        
                        workbookSum = load_workbook(filenameSum)
                        workbookSum.active

                        InputDataToSum(workbookSum,'Sheet1','B',str(ID))
                        InputDataToSum(workbookSum,'Sheet1','C',self.x2_value)
                        InputDataToSum(workbookSum,'Sheet1','D',self.x3_value)
                        InputDataToSum(workbookSum,'Sheet1','E',self.x4_value)
                        if budgetChoice == "งบประมาณแผ่นดิน":
                            InputDataToSum(workbookSum,'Sheet1','F',self.sum) 
                            InputDataToSum(workbookSum,'Sheet1','G',0) 
                            messagebox.showinfo("Success","ส่งข้อมูลเรียบร้อย")
                        elif budgetChoice == "งบประมาณรายได้":
                            InputDataToSum(workbookSum,'Sheet1','G',self.sum) 
                            InputDataToSum(workbookSum,'Sheet1','F',0) 
                            messagebox.showinfo("Success","ส่งข้อมูลเรียบร้อย")
                        else:
                            messagebox.showinfo("INVALID", "กรุณาเลือกงบ")
                        workbookSum.save(filenameSum)


                def deletee():
                    messagebox.showinfo("Delete","ล้างข้อมูลเรียบร้อย")
                    self.entry_2.delete(0,END)
                    entry_3.delete(0,END)
                    entry_4.delete(0,END)
                    entry_5.delete(0,END)
                    entry_6.delete(0,END)
                    entry_7.delete(0,END)
                    entry_8.delete(0,END)
                    entry_9.delete(0,END)
                    entry_10.delete(0,END)
                    entry_11.delete(0,END)
                    entry_12.delete(0,END)
                    entry_13.delete(0,END)
                    entry_14.delete(0,END)
                    entry_15.delete(0,END)
                    entry_16.delete(0,END)
                
                def close():
                    window2.destroy()
                    

                button1 = Button(window2,text = " ส่งข้อมูล ",font=("Inter SemiBold", 15 * -1),command = success).place(x=70,y=620)
                button2 = Button(window2,text = " ล้างข้อมูล ",font=("Inter SemiBold", 15 * -1),command = deletee).place(x=180,y=620)
                button3 = Button(window2,text = " ปิด ",font=("Inter SemiBold", 15 * -1),command = close).place(x=300,y=620)
            else:
                messagebox.showinfo("INVALID", "โปรดใส่รหัสวิชาให้ถูกต้อง")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")


