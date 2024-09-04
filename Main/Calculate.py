from openpyxl import load_workbook

#สำหรับหาแบบรวมทุกรายวิชา
def monthCal(workbook, month,year, income):
    wb = load_workbook(workbook)
    ws = wb.active
    num = 4
    result_dict = {}  # Initialize an empty dictionary to store the results
    while num <= ws.max_row:
        Month = ws['C' + str(num)].value
        Year = ws['C' + str(num)].value
        if Month and Year:
            Month_value = Month.split('/')
            Year_value = Year.split('/')
            if len(Year_value) >= 3 and int(Year_value[2]) == int(year):
                if len(Month_value) >= 2 and int(Month_value[1]) == int(month):
                    if income == 'งบประมาณแผ่นดิน':
                        category = ws['E' + str(num)].value  # Assuming category is in column D
                        cell_value = ws['F' + str(num)].value
                    elif income == 'งบประมาณรายได้':
                        category = ws['E' + str(num)].value  # Assuming category is in column E
                        cell_value = ws['G' + str(num)].value
                    else:
                        category = None  # Handle other income types here
                        cell_value = None
                    if category is not None and cell_value is not None:
                        if category not in result_dict:
                            result_dict[category] = cell_value
                        else:
                            result_dict[category] += cell_value

                # print(f"Category: {category}, Cell Value: {cell_value}")
                # print(f"Result Dictionary: {result_dict}")
        num += 1
    return result_dict

def yearCal(workbook, year, income):
    wb = load_workbook(workbook)
    ws = wb.active
    num = 4
    result_dict = {}
    # Renamed "sum" to "total_sum" to avoid shadowing the built-in sum() function
    while num <= ws.max_row:  # Use ws.max_row to limit the loop to valid rows
        Year = ws['C' + str(num)].value
        if Year:
            Year_value = Year.split('/')
            if len(Year_value) >= 3 and int(Year_value[2]) == int(year):
                if income == 'งบประมาณแผ่นดิน':
                    category = ws['E' + str(num)].value
                    cell_value = ws['F' + str(num)].value
                elif income == 'งบประมาณรายได้':
                    category = ws['E' + str(num)].value
                    cell_value = ws['G' + str(num)].value
                else:
                    category = None
                    cell_value = None  # Handle other income types here
                if category is not None and cell_value is not None:
                    if category not in result_dict:
                        result_dict[category] = cell_value
                    else:
                        result_dict[category] += cell_value # Exit the loop if the cell value is None

                #print(f"Cell Value: {cell_value}")
                #print(f"Total Sum: {total_sum}")
        num += 1  # Move to the next row
    return result_dict

#สำหรับหาแบบแยกรายวิชา
def SubjectMCal(workbook,month,year,income,ID):
    wb = load_workbook(workbook)
    ws = wb.active
    num = 4
    result_dict = {}  
    while num <= ws.max_row:  
        Month = ws['C' + str(num)].value
        Year = ws['C' + str(num)].value
        SubjectID = ws['B' + str(num)].value
        if Month:
            Month_value = Month.split('/')
            Year_value = Year.split('/')
            if len(Year_value) >= 3 and int(Year_value[2]) == int(year) and int(SubjectID) == int(ID):
                if len(Month_value) >= 2 and int(Month_value[1]) == int(month) and int(SubjectID) == int(ID):
                    if income == 'งบประมาณแผ่นดิน':
                        category = ws['E' + str(num)].value
                        cell_value = ws['F' + str(num)].value
                    elif income == 'งบประมาณรายได้':
                        category = ws['E' + str(num)].value
                        cell_value = ws['G' + str(num)].value
                    else:
                        category = None
                        cell_value = None  # Handle other income types here
                    if category not in result_dict:
                        if category not in result_dict:
                            result_dict[category] = cell_value
                        else:
                            result_dict[category] += cell_value # Exit the loop if the cell value is None

                    #print(f"Cell Value: {cell_value}")
                    #print(f"Total Sum: {total_sum}")
        num += 1  # Move to the next row
    return result_dict
        

def SubjectYCal(workbook, year, income,ID):
    wb = load_workbook(workbook)
    ws = wb.active
    num = 4
    result_dict = {}  # Renamed "sum" to "total_sum" to avoid shadowing the built-in sum() function
    while num <= ws.max_row:  # Use ws.max_row to limit the loop to valid rows
        Year = ws['C' + str(num)].value
        SubjectID = ws['B' + str(num)].value
        if Year:
            Year_value = Year.split('/')
            if len(Year_value) >= 3 and int(Year_value[2]) == int(year) and int(SubjectID) == int(ID):
                if income == 'งบประมาณแผ่นดิน':
                    category = ws['E' + str(num)].value
                    cell_value = ws['F' + str(num)].value
                elif income == 'งบประมาณรายได้':
                    category = ws['E' + str(num)].value
                    cell_value = ws['G' + str(num)].value
                else:
                    category = None
                    cell_value = None  # Handle other income types here
                if category not in result_dict:
                    if category not in result_dict:
                        result_dict[category] = cell_value
                    else:
                        result_dict[category] += cell_value # Exit the loop if the cell value is None

                #print(f"Cell Value: {cell_value}")
                #print(f"Total Sum: {total_sum}")
        num += 1  # Move to the next row
    return result_dict

