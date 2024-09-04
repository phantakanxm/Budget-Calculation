from openpyxl import load_workbook
filename = 'Test01.xlsx'
sheet_name = '501004'
wb = load_workbook(filename)
sheet = wb[sheet_name]

MergeCell = sheet.merged_cells.ranges
print(MergeCell)
cell = sheet['D4']

def testMerge(row, column):
    cell = sheet.cell(row, column)
    for mergedCell in sheet.merged_cells.ranges:
        if (cell.coordinate in mergedCell):
            return True
    return False
print(testMerge(4,4))

if testMerge(4,4):
    print(True)
else:
    print(False)

