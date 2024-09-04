from openpyxl import load_workbook
import sys

def columnSum(workbook, sheet_name, column):
    wb = load_workbook(workbook)
    ws = wb[sheet_name]

    total_sum = 0
    for row in range(7, ws.max_row + 1):
        cell = ws[column + str(row)].value
        print(cell)
        total_sum += cell        

    wb.close()
    return total_sum

filename = 'Test01.xlsx'
s = columnSum(filename, '501004', 'E')
print(s)
